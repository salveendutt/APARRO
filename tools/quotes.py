import openpyxl

# Function to add double quotes around each cell's content in an XLSX file
def add_quotes_to_xlsx(input_file, output_file):
    # Load the input workbook
    workbook = openpyxl.load_workbook(input_file)
    
    # Create an output workbook
    output_workbook = openpyxl.Workbook()
    
    for sheet_name in workbook.sheetnames:
        # Get the current sheet from the input workbook
        sheet = workbook[sheet_name]
        
        # Create a new sheet with the same name in the output workbook
        output_sheet = output_workbook.create_sheet(title=sheet_name)
        
        # Iterate through rows and columns, adding double quotes to cell content
        for row in sheet.iter_rows():
            output_row = []
            for cell in row:
                quoted_content = f'"{cell.value}"' if cell.value is not None else ""
                output_row.append(quoted_content)
            output_sheet.append(output_row)

    # Remove the default sheet created by openpyxl
    output_workbook.remove(output_workbook['Sheet'])
    
    # Save the modified XLSX to the output file
    output_workbook.save(output_file)

if __name__ == "__main__":
    input_file = "Datasets/order_dataset_1.0.xlsx"  # Replace with your input XLSX file path
    output_file = "Datasets/order_dataset_1.1.xlsx"  # Replace with your desired output XLSX file path

    add_quotes_to_xlsx(input_file, output_file)
    print(f'Quotes added to {input_file} and saved as {output_file}')
